
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

wb_obj = openpyxl.load_workbook('selectkaro.xlsx')
sheet = wb_obj.active

for i in range(2, sheet.max_row+1):
    username = sheet.cell(row=i, column=1)
    password = sheet.cell(row=i, column=2)
    amount = sheet.cell(row=i, column=3)
    amz_wallet = sheet.cell(row=i, column=4)
    amz_mail = sheet.cell(row=i, column=5)
    amz_password = sheet.cell(row=i, column=6)
    status = sheet.cell(row=i, column=7)

    try:
        driver = webdriver.Chrome()
        driver.get("https://selectkaro.com/")
        driver.maximize_window()

        profile = driver.find_element(By.XPATH,"//button[contains(@class,'nav-item nav-link btn fas fa-user-circle fa-2x nav_text har2 btnlogin')]")
        profile.click()
        time.sleep(2)

        enter_email = driver.find_element(By.XPATH, "//input[contains(@name,'Email')]")
        enter_email.send_keys(username.value)
        time.sleep(2)

        enter_password = driver.find_element(By.XPATH, "//input[contains(@name,'Password')]")
        enter_password.send_keys(password.value)
        time.sleep(2)

        login = driver.find_element(By.XPATH, "//input[contains(@class,'btn orange_item btn-block')]")
        login.click()
        time.sleep(3)

        wallet = driver.find_element(By.XPATH, "//a[contains(@href,'/makepayment')]")
        wallet.click()
        driver.execute_script("window.scrollTo(0,400)")
        time.sleep(2)

        payu = driver.find_element(By.XPATH, "//label[contains(@for,'payment-PayU')]")
        payu.click()
        time.sleep(2)

        enter_amount = driver.find_element(By.XPATH, "//input[contains(@name,'Amount')]")
        enter_amount.send_keys(amount.value)
        time.sleep(1)

        proceed = driver.find_element(By.XPATH, "//button[contains(@class,'btn btn-rzee-blue wow flipInY  animated')]")
        proceed.click()
        time.sleep(3)

        wallets = driver.find_element(By.XPATH, "//div[contains(@class,'ewallets pull-left')]")
        wallets.click()
        time.sleep(2)

        select_wallets = driver.find_element(By.XPATH, "//select[contains(@id,'externalWalletsDrop')]")
        select_wallets.click()
        time.sleep(2)
        select_wallets.send_keys(amz_wallet.value)
        time.sleep(2)

        paynow = driver.find_element(By.XPATH, "//button[contains(@id,'citrusEwalletsButton')]")
        paynow.click()
        time.sleep(2)

        amazon_email = driver.find_element(By.XPATH, "//input[contains(@id,'ap_email')]")
        amazon_email.send_keys(amz_mail.value)
        time.sleep(2)

        amazon_password = driver.find_element(By.XPATH, "//input[contains(@id,'ap_password')]")
        amazon_password.send_keys(amz_password.value)
        time.sleep(2)

        amz_signin = driver.find_element(By.XPATH, "//input[contains(@id,'signInSubmit')]")
        amz_signin.click()
        time.sleep(2)

        amz_paynow = driver.find_element(By.XPATH, "//input[contains(@name,'ppw-widgetEvent:SetPaymentPlanSelectContinueEvent')]")
        amz_paynow.click()
        time.sleep(6)

        status.value = "Success"
        wb_obj.save("selectkaro.xlsx")

        driver.close()

    except:
        print("=======>>> Something went wrong  <<<=======")

        status.value = "Success"
        wb_obj.save("selectkaro.xlsx")

        driver.close()
        continue
